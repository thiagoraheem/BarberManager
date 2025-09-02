import io
import pandas as pd
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from io import BytesIO
import base64

from models import Appointment, Client, Service, Sale, User, AppointmentStatus, PaymentMethod
from crud import get_appointments, get_clients, get_services, get_sales

class ReportGenerator:
    """Advanced report generator with Excel and PDF support"""
    
    def __init__(self, db: Session):
        self.db = db
        self.styles = getSampleStyleSheet()
        
    def generate_financial_report(
        self, 
        start_date: date, 
        end_date: date,
        format_type: str = "excel"
    ) -> bytes:
        """Generate comprehensive financial report"""
        
        # Get sales data
        sales = self.db.query(Sale).filter(
            func.date(Sale.criado_em) >= start_date,
            func.date(Sale.criado_em) <= end_date
        ).all()
        
        # Get appointments data
        appointments = self.db.query(Appointment).filter(
            func.date(Appointment.data_hora) >= start_date,
            func.date(Appointment.data_hora) <= end_date
        ).all()
        
        if format_type == "excel":
            return self._generate_financial_excel(sales, appointments, start_date, end_date)
        else:
            return self._generate_financial_pdf(sales, appointments, start_date, end_date)
    
    def _generate_financial_excel(self, sales, appointments, start_date, end_date) -> bytes:
        """Generate financial report in Excel format"""
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("Resumo Financeiro")
        self._create_financial_summary_sheet(ws_summary, sales, appointments, start_date, end_date)
        
        # 2. Sales by Day
        ws_daily = wb.create_sheet("Vendas Diárias")
        self._create_daily_sales_sheet(ws_daily, sales, start_date, end_date)
        
        # 3. Payment Methods
        ws_payment = wb.create_sheet("Métodos de Pagamento")
        self._create_payment_methods_sheet(ws_payment, sales)
        
        # 4. Services Performance
        ws_services = wb.create_sheet("Performance Serviços")
        self._create_services_performance_sheet(ws_services, sales, appointments)
        
        # 5. Barber Performance
        ws_barbers = wb.create_sheet("Performance Barbeiros")
        self._create_barber_performance_sheet(ws_barbers, appointments)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_financial_summary_sheet(self, ws, sales, appointments, start_date, end_date):
        """Create financial summary sheet"""
        
        # Title
        ws['A1'] = "RELATÓRIO FINANCEIRO"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        
        # Total revenue
        total_revenue = sum(sale.total for sale in sales)
        total_appointments = len(appointments)
        
        # Revenue by payment method
        revenue_cash = sum(sale.total for sale in sales if sale.metodo_pagamento == PaymentMethod.DINHEIRO)
        revenue_card = sum(sale.total for sale in sales if sale.metodo_pagamento in [PaymentMethod.CARTAO_CREDITO, PaymentMethod.CARTAO_DEBITO])
        revenue_pix = sum(sale.total for sale in sales if sale.metodo_pagamento == PaymentMethod.PIX)
        
        # Summary data
        summary_data = [
            ["Métrica", "Valor"],
            ["Total de Vendas", f"R$ {total_revenue:.2f}"],
            ["Total de Agendamentos", total_appointments],
            ["Ticket Médio", f"R$ {total_revenue/len(sales) if sales else 0:.2f}"],
            ["", ""],
            ["Vendas por Método de Pagamento", ""],
            ["Dinheiro", f"R$ {revenue_cash:.2f}"],
            ["Cartão", f"R$ {revenue_card:.2f}"],
            ["PIX", f"R$ {revenue_pix:.2f}"],
        ]
        
        # Write data
        for row_idx, row_data in enumerate(summary_data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_daily_sales_sheet(self, ws, sales, start_date, end_date):
        """Create daily sales analysis sheet"""
        
        # Group sales by date
        daily_sales = {}
        for sale in sales:
            sale_date = sale.criado_em.date()
            if sale_date not in daily_sales:
                daily_sales[sale_date] = 0
            daily_sales[sale_date] += sale.total
        
        # Create date range
        current_date = start_date
        headers = ["Data", "Vendas (R$)", "Dia da Semana"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add daily data
        while current_date <= end_date:
            day_sales = daily_sales.get(current_date, 0)
            weekday = current_date.strftime("%A")
            weekday_pt = {
                "Monday": "Segunda", "Tuesday": "Terça", "Wednesday": "Quarta",
                "Thursday": "Quinta", "Friday": "Sexta", "Saturday": "Sábado", "Sunday": "Domingo"
            }.get(weekday, weekday)
            
            ws.append([
                current_date.strftime("%d/%m/%Y"),
                f"{day_sales:.2f}",
                weekday_pt
            ])
            current_date += timedelta(days=1)
        
        # Add chart
        chart = BarChart()
        chart.title = "Vendas Diárias"
        chart.y_axis.title = "Valor (R$)"
        chart.x_axis.title = "Data"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, "E2")
    
    def _create_payment_methods_sheet(self, ws, sales):
        """Create payment methods analysis sheet"""
        
        # Group by payment method
        payment_data = {}
        for sale in sales:
            method = sale.metodo_pagamento.value
            if method not in payment_data:
                payment_data[method] = {"count": 0, "total": 0}
            payment_data[method]["count"] += 1
            payment_data[method]["total"] += sale.total
        
        # Headers
        headers = ["Método de Pagamento", "Quantidade", "Total (R$)", "Percentual"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Calculate total for percentage
        total_sales = sum(data["total"] for data in payment_data.values())
        
        # Add data
        for method, data in payment_data.items():
            percentage = (data["total"] / total_sales * 100) if total_sales > 0 else 0
            ws.append([
                method,
                data["count"],
                f"{data['total']:.2f}",
                f"{percentage:.1f}%"
            ])
    
    def _create_services_performance_sheet(self, ws, sales, appointments):
        """Create services performance analysis sheet"""
        
        # Group by service
        service_data = {}
        
        # From sales
        for sale in sales:
            for item in sale.itens:
                service_name = item.servico.nome if hasattr(item, 'servico') else f"Serviço ID {item.servico_id}"
                if service_name not in service_data:
                    service_data[service_name] = {"sales_count": 0, "sales_total": 0, "appointments": 0}
                service_data[service_name]["sales_count"] += item.quantidade
                service_data[service_name]["sales_total"] += item.subtotal
        
        # From appointments
        for appointment in appointments:
            service_name = appointment.servico.nome
            if service_name not in service_data:
                service_data[service_name] = {"sales_count": 0, "sales_total": 0, "appointments": 0}
            service_data[service_name]["appointments"] += 1
        
        # Headers
        headers = ["Serviço", "Vendas Qtd", "Vendas Total (R$)", "Agendamentos", "Ticket Médio"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for service, data in sorted(service_data.items(), key=lambda x: x[1]["sales_total"], reverse=True):
            avg_ticket = data["sales_total"] / data["sales_count"] if data["sales_count"] > 0 else 0
            ws.append([
                service,
                data["sales_count"],
                f"{data['sales_total']:.2f}",
                data["appointments"],
                f"{avg_ticket:.2f}"
            ])
    
    def _create_barber_performance_sheet(self, ws, appointments):
        """Create barber performance analysis sheet"""
        
        # Group by barber
        barber_data = {}
        for appointment in appointments:
            barber_name = appointment.barbeiro.nome
            if barber_name not in barber_data:
                barber_data[barber_name] = {
                    "appointments": 0,
                    "completed": 0,
                    "cancelled": 0,
                    "revenue": 0
                }
            
            barber_data[barber_name]["appointments"] += 1
            
            if appointment.status == AppointmentStatus.CONCLUIDO:
                barber_data[barber_name]["completed"] += 1
                barber_data[barber_name]["revenue"] += appointment.servico.preco
            elif appointment.status == AppointmentStatus.CANCELADO:
                barber_data[barber_name]["cancelled"] += 1
        
        # Headers
        headers = ["Barbeiro", "Total Agendamentos", "Concluídos", "Cancelados", "Taxa Sucesso", "Receita (R$)"]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for barber, data in sorted(barber_data.items(), key=lambda x: x[1]["revenue"], reverse=True):
            success_rate = (data["completed"] / data["appointments"] * 100) if data["appointments"] > 0 else 0
            ws.append([
                barber,
                data["appointments"],
                data["completed"],
                data["cancelled"],
                f"{success_rate:.1f}%",
                f"{data['revenue']:.2f}"
            ])
    
    def _generate_financial_pdf(self, sales, appointments, start_date, end_date) -> bytes:
        """Generate financial report in PDF format"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        story.append(Paragraph("RELATÓRIO FINANCEIRO", title_style))
        story.append(Paragraph(f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}", self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Financial Summary
        total_revenue = sum(sale.total for sale in sales)
        total_appointments = len(appointments)
        avg_ticket = total_revenue / len(sales) if sales else 0
        
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Vendas', f'R$ {total_revenue:.2f}'],
            ['Total de Agendamentos', str(total_appointments)],
            ['Ticket Médio', f'R$ {avg_ticket:.2f}'],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Payment Methods Analysis
        story.append(Paragraph("Análise por Método de Pagamento", self.styles['Heading2']))
        
        payment_data = {}
        for sale in sales:
            method = sale.metodo_pagamento.value
            if method not in payment_data:
                payment_data[method] = {"count": 0, "total": 0}
            payment_data[method]["count"] += 1
            payment_data[method]["total"] += sale.total
        
        payment_table_data = [['Método', 'Quantidade', 'Total (R$)', 'Percentual']]
        total_sales = sum(data["total"] for data in payment_data.values())
        
        for method, data in payment_data.items():
            percentage = (data["total"] / total_sales * 100) if total_sales > 0 else 0
            payment_table_data.append([
                method,
                str(data["count"]),
                f'R$ {data["total"]:.2f}',
                f'{percentage:.1f}%'
            ])
        
        payment_table = Table(payment_table_data)
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(payment_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_client_report(self, format_type: str = "excel") -> bytes:
        """Generate client analysis report"""
        
        clients = self.db.query(Client).filter(Client.ativo == True).all()
        
        if format_type == "excel":
            return self._generate_client_excel(clients)
        else:
            return self._generate_client_pdf(clients)
    
    def _generate_client_excel(self, clients) -> bytes:
        """Generate client report in Excel format"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Clientes"
        
        # Headers
        headers = [
            "Nome", "Email", "Telefone", "Data Cadastro", 
            "LGPD Aceito", "Total Agendamentos", "Último Agendamento"
        ]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add client data
        for client in clients:
            appointments_count = len(client.agendamentos)
            last_appointment = max(client.agendamentos, key=lambda x: x.data_hora).data_hora if client.agendamentos else None
            
            ws.append([
                client.nome,
                client.email or "",
                client.telefone,
                client.criado_em.strftime("%d/%m/%Y"),
                "Sim" if client.aceite_lgpd else "Não",
                appointments_count,
                last_appointment.strftime("%d/%m/%Y") if last_appointment else ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _generate_client_pdf(self, clients) -> bytes:
        """Generate client report in PDF format"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        
        story.append(Paragraph("RELATÓRIO DE CLIENTES", title_style))
        story.append(Spacer(1, 20))
        
        # Summary
        total_clients = len(clients)
        lgpd_compliant = sum(1 for client in clients if client.aceite_lgpd)
        
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Clientes Ativos', str(total_clients)],
            ['Clientes com LGPD', str(lgpd_compliant)],
            ['Taxa de Conformidade LGPD', f'{(lgpd_compliant/total_clients*100):.1f}%' if total_clients > 0 else '0%'],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Client list (first 50)
        story.append(Paragraph("Lista de Clientes (Primeiros 50)", self.styles['Heading2']))
        
        client_data = [['Nome', 'Email', 'Telefone', 'LGPD']]
        for client in clients[:50]:
            client_data.append([
                client.nome,
                client.email or "N/A",
                client.telefone,
                "✓" if client.aceite_lgpd else "✗"
            ])
        
        client_table = Table(client_data)
        client_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8)
        ]))
        
        story.append(client_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_appointment_report(
        self, 
        start_date: date, 
        end_date: date, 
        format_type: str = "excel"
    ) -> bytes:
        """Generate appointment analysis report"""
        
        appointments = self.db.query(Appointment).filter(
            func.date(Appointment.data_hora) >= start_date,
            func.date(Appointment.data_hora) <= end_date
        ).all()
        
        if format_type == "excel":
            return self._generate_appointment_excel(appointments, start_date, end_date)
        else:
            return self._generate_appointment_pdf(appointments, start_date, end_date)
    
    def _generate_appointment_excel(self, appointments, start_date, end_date) -> bytes:
        """Generate appointment report in Excel format"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Agendamentos"
        
        # Title
        ws['A1'] = "RELATÓRIO DE AGENDAMENTOS"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        
        # Headers
        headers = [
            "Data/Hora", "Cliente", "Barbeiro", "Serviço", 
            "Status", "Valor", "Observações"
        ]
        ws.append([""] * len(headers))  # Empty row
        ws.append(headers)
        
        # Style header
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add appointment data
        for appointment in appointments:
            ws.append([
                appointment.data_hora.strftime("%d/%m/%Y %H:%M"),
                appointment.cliente.nome,
                appointment.barbeiro.nome,
                appointment.servico.nome,
                appointment.status.value,
                f"R$ {appointment.servico.preco:.2f}",
                appointment.observacoes or ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _generate_appointment_pdf(self, appointments, start_date, end_date) -> bytes:
        """Generate appointment report in PDF format"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        
        story.append(Paragraph("RELATÓRIO DE AGENDAMENTOS", title_style))
        story.append(Paragraph(f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}", self.styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary by status
        status_count = {}
        for appointment in appointments:
            status = appointment.status.value
            status_count[status] = status_count.get(status, 0) + 1
        
        summary_data = [['Status', 'Quantidade']]
        for status, count in status_count.items():
            summary_data.append([status, str(count)])
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

# Utility function to create report generator
def create_report_generator(db: Session) -> ReportGenerator:
    """Create a new report generator instance"""
    return ReportGenerator(db)